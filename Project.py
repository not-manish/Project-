from flask import Flask, request, render_template_string, send_file, abort, redirect, session
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import io
import uuid
import tempfile
import os
app = Flask(__name__)
app.secret_key = "final-stable-session-key-2025"
DATA_STORE = {}
from openpyxl import load_workbook
import re
def clean(x):
    if x is None:
        return ""
    if isinstance(x, datetime):
        return x.strftime("%b %d, %Y")
    return str(x).strip().replace("\u00a0", " ").replace("\n", " ")
CLASS_INCHARGE_FORM = '''
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=5" />
<title>Set Class Incharge</title>
<style>
    :root{
      --card:#ffffff;
      --muted:#64748b;
      --accent:#8b5cf6;
    }
    *{box-sizing:border-box}
    body{
      margin:0;
      font-family: -apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      background: #0f172a;
      color:#1e293b;
      padding:20px 16px;
      display:flex;
      justify-content:center;
      min-height:100vh;
    }
    .card{width:100%;max-width:600px;background:var(--card);border-radius:16px;padding:32px 24px;box-shadow:0 4px 12px rgba(0,0,0,0.06)}
    h2{color:var(--accent);margin:0 0 16px 0;font-size:24px}
    .info{color:var(--muted);margin:12px 0 24px 0;font-size:14px;line-height:1.5}
    label{display:block;font-weight:700;margin-top:16px;color:#334155;font-size:14px}
    input{width:100%;padding:16px;border-radius:10px;border:2px solid #e2e8f0;margin-top:8px;font-size:16px;font-family:inherit}
    input:focus{border-color:var(--accent);outline:none;box-shadow:0 0 0 3px rgba(139,92,246,0.1)}
    .btn{margin-top:24px;padding:16px;background:var(--accent);color:#fff;border:0;border-radius:10px;font-weight:800;
     width:100%;font-size:16px;cursor:pointer;touch-action:manipulation}
    .btn:hover{opacity:0.9}
    .btn:active{transform:scale(0.98)}
    .back-btn{display:inline-block;margin-bottom:16px;padding:10px 14px;border-radius:10px;
    background:#e2e8f0;color:#0f172a;border:0;font-weight:700;cursor:pointer;font-size:13px;touch-action:manipulation}
    .back-btn:hover{background:#cbd5e1}
    @media (max-width:480px){
      body{padding:16px 12px}
      .card{padding:20px 14px;border-radius:12px}
      h2{font-size:20px;margin-bottom:10px}
      label{font-size:12px;margin-top:12px}
      input{padding:12px;font-size:16px}
      .btn{padding:10px 12px;font-size:13px;margin-top:16px}
      .back-btn{padding:8px 10px;font-size:11px}
    }
</style>
</head>
<body>
<div class="card">
<button class="back-btn" onclick="window.location='/bulk'">← Back to Bulk</button>
<h2>Set Class Incharge Name</h2>
<p class="info">This name will be used for all {{ count }} certificates in this batch.</p>
<form method="post" action="/set-incharge">
  <label>Class Incharge Full Name</label>
  <input name="class_incharge_name" type="text" placeholder="e.g., Ms. Priya Sharma" maxlength="50" required autofocus>
  <button class="btn" type="submit">Continue to First Student</button>
</form>
</div>
</body>
</html>
'''
BULK_UPLOAD_HTML = '''
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=5" />
  <title>Bulk Certificate Upload</title>
  <style>
    :root{
      --card:#ffffff;
      --muted:#64748b;
      --accent:#8b5cf6;
    }
    *{box-sizing:border-box}
    body{
      margin:0;
      font-family: -apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      background: #0f172a;
      color:#1e293b;
      -webkit-font-smoothing:antialiased;
      -moz-osx-font-smoothing:grayscale;
      padding:20px 16px;
      display:flex;
      justify-content:center;
      min-height:100vh;
    }
    .card{
      width:100%;
      max-width:900px;
      background:var(--card);
      border-radius:16px;
      box-shadow:0 4px 12px rgba(0,0,0,0.08);
      padding:32px 24px;
    }
    h1{font-size:28px; margin:0 0 8px 0; color:var(--accent)}
    .muted{color:var(--muted); margin-top:6px;font-size:14px}
    input[type=file]{width:100%; padding:14px; border-radius:10px; border:2px solid #e2e8f0; background:#fff;font-size:16px}
    input[type=file]:focus{border-color:var(--accent);outline:none}
    .btn{display:inline-block;margin-top:16px;padding:14px 20px;border-radius:10px;background:var(--accent);
     color:#fff;border:0;font-weight:800;cursor:pointer;font-size:15px;touch-action:manipulation}
    .btn:active{transform:scale(0.98)}
    .back-btn{display:inline-block;margin-bottom:16px;padding:10px 14px;border-radius:10px;background:#e2e8f0;
     color:#0f172a;border:0;font-weight:700;cursor:pointer;font-size:13px;touch-action:manipulation}
    .back-btn:hover{background:#cbd5e1}
    @media (max-width:768px){ .card{padding:24px 20px} h1{font-size:24px}}
    @media (max-width:480px){ 
      body{padding:16px 12px}
      .card{padding:18px 14px;border-radius:12px} 
      h1{font-size:22px;margin-bottom:6px}
      .muted{font-size:12px;margin-top:3px}
      input[type=file]{padding:11px;font-size:13px}
      .btn{padding:10px 12px;font-size:13px;margin-top:12px}
      .back-btn{padding:8px 10px;font-size:11px}
    }
  </style>
</head>
<body>
  <div class="card">
    <button class="back-btn" onclick="window.location='/'">← Back to Main</button>
    <h1>Bulk Upload</h1>
    <div class="muted">Upload Haryana MIS Excel (first sheet will be processed)</div>
    <form method="post" enctype="multipart/form-data" style="margin-top:18px">
      <input type="file" name="excel" required>
      <button class="btn" type="submit">Upload & Start</button>
    </form>
  </div>
</body>
</html>
'''
BULK_FORM_HTML = '''
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=5" />
<title>Generate Character Certificate</title>
<style>
    :root{
      --card:#ffffff;
      --muted:#64748b;
      --accent:#8b5cf6;
    }
    *{box-sizing:border-box}
    body{
      margin:0;
      font-family: -apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      background: #0f172a;
      color:#1e293b;
      padding:20px 16px;
      display:flex;
      justify-content:center;
      min-height:100vh;
    }
    .card{width:100%;max-width:720px;background:var(--card);border-radius:16px;padding:24px;box-shadow:0 4px 12px rgba(0,0,0,0.06)}
    h2{color:var(--accent);margin:0 0 6px 0;font-size:22px}
    .info{color:var(--muted);margin:6px 0;font-size:13px;line-height:1.4}
    label{display:block;font-weight:700;margin-top:14px;color:#334155;font-size:13px}
    input{width:100%;padding:14px;border-radius:10px;border:2px solid #e2e8f0;margin-top:6px;font-size:16px;font-family:inherit}
    input:focus{border-color:var(--accent);outline:none;box-shadow:0 0 0 3px rgba(139,92,246,0.1)}
    .row-inputs input{display:inline-block;width:32%;margin-right:2%}
    .row-inputs input:last-child{margin-right:0}
    .btn{margin-top:18px;padding:14px;background:var(--accent);color:#fff;border:0;border-radius:10px;
     font-weight:800;width:100%;cursor:pointer;font-size:15px;touch-action:manipulation}
    .btn:active{transform:scale(0.98)}
    .back-btn{display:inline-block;margin-bottom:12px;padding:10px 14px;border-radius:10px;background:#e2e8f0;
     color:#0f172a;border:0;font-weight:700;cursor:pointer;font-size:12px;touch-action:manipulation}
    .back-btn:hover{background:#cbd5e1}
    .date-input{width:100%;padding:14px;border-radius:10px;border:2px solid #e2e8f0;margin-top:6px;font-size:16px}
    .date-input-group input { width: 32%; margin-right: 2%; }
    .date-input-group input:last-child { margin-right: 0; }
    .readonly{background:#f1f5f9;color:#64748b}
    .credit{
      font-size:12px;
      margin-top:16px;
      text-align:center;
      font-weight:700;
      letter-spacing:1.2px;
      padding:10px 18px;
      border-radius:10px;
      background:linear-gradient(135deg,#7c3aed 0%,#a855f7 25%,#d946ef 50%,#ec4899 75%,#f43f5e 100%);
      background-size:200% 200%;
      -webkit-background-clip:text;
      -webkit-text-fill-color:transparent;
      background-clip:text;
      position:relative;
      transition:all 0.35s cubic-bezier(0.34,1.56,0.64,1);
      box-shadow:0 2px 8px rgba(139,92,246,0.15),0 0 12px rgba(217,70,239,0.08),inset 0 1px 2px rgba(255,255,255,0.15);
      display:inline-block;
    }
    .credit::before{
      content:'✨';
      position:absolute;
      left:0;
      opacity:0;
      transition:all 0.35s cubic-bezier(0.34,1.56,0.64,1);
      transform:translateX(-8px) scale(0.8);
      font-size:11px;
    }
    .credit::after{
      content:'✨';
      position:absolute;
      right:0;
      opacity:0;
      transition:all 0.35s cubic-bezier(0.34,1.56,0.64,1);
      transform:translateX(8px) scale(0.8);
      font-size:11px;
    }
    .credit:hover{
      transform:translateY(-4px) scale(1.04) rotateX(2deg);
      background-position:100% 100%;
      box-shadow:0 8px 20px rgba(139,92,246,0.35),0 0 30px rgba(217,70,239,0.25),
      0 0 40px rgba(244,63,94,0.1),inset 0 1px 2px rgba(255,255,255,0.25);
    }
    .credit:hover::before{
      opacity:1;
      transform:translateX(-15px) scale(1);
      animation:starFloat 0.6s ease-out;
    }
    .credit:hover::after{
      opacity:1;
      transform:translateX(15px) scale(1);
      animation:starFloat 0.6s ease-out;
    }
    @keyframes starFloat{
      0%{transform:translateX(0) scale(0.8) translateY(0)}
      50%{transform:translateX(0) scale(1.1) translateY(-3px)}
      100%{transform:translateX(0) scale(1) translateY(0)}
    }
    @media (max-width:768px){ .card{padding:20px} }
    @media (max-width:640px){ 
      .row-inputs input{width:100%;display:block;margin-right:0;margin-bottom:6px}
      .date-input-group input{width:100%;display:block;margin-right:0;margin-bottom:6px}
    }
    @media (max-width:480px){
      body{padding:16px 12px}
      .card{padding:14px 10px;border-radius:12px}
      h2{font-size:19px}
      .info{font-size:12px}
      label{font-size:11px;margin-top:10px}
      input{padding:11px;font-size:16px;margin-top:3px}
      .btn{padding:10px 11px;font-size:12px;margin-top:12px}
      .back-btn{padding:8px 9px;font-size:10px;margin-bottom:8px}
    }
</style>
</head>
<body>
<div class="card">
<button class="back-btn" onclick="window.location='/bulk'">← Back to Bulk</button>
<h2>Generate Character Certificate</h2>
<p><b>Student Name:</b> {{ s.student_name }}</p>
<p><b>Father Name:</b> {{ s.father_name }}</p>
<p><b>Mother Name:</b> {{ s.mother_name }}</p>
<p class="info"><b>Date of Birth:</b> {{ s.dob_display }}</p>
<hr>
<form method="post" action="/generate-one">
  <label>Roll Number (Max 30 characters)</label>
  <input name="roll_number" type="text" placeholder="Enter Roll Number" maxlength="30" required>
  <label>Date of Admission (DD/MM/YYYY)</label>
  <div class="row-inputs date-input-group">
    <input name="doa_day" type="number" min="1" max="31" placeholder="DD (1-31)" maxlength="2" value="{{ s.doa_day }}">
    <input name="doa_month" type="number" min="1" max="12" placeholder="MM (1-12)" maxlength="2" value="{{ s.doa_month }}">
    <input name="doa_year" type="number" min="1900" max="2100" placeholder="YYYY (e.g. 2025)" maxlength="4" value="{{ s.doa_year }}">
  </div>
  <label>Subjects (Max 30 characters each)</label>
  <div class="row-inputs">
    <input name="subject1" type="text" placeholder="Subject 1" maxlength="30">
    <input name="subject2" type="text" placeholder="Subject 2" maxlength="30">
    <input name="subject3" type="text" placeholder="Subject 3" maxlength="30">
  </div>
  <div class="row-inputs">
    <input name="subject4" type="text" placeholder="Subject 4" maxlength="30">
    <input name="subject5" type="text" placeholder="Subject 5" maxlength="30">
    <input name="subject6" type="text" placeholder="Subject 6" maxlength="30">
  </div>
  <label>Marks Obtained</label>
  <input name="marks_obtained" value="{{ s.get('marks_obtained','') }}" type="number" placeholder="Numeric value" required>
  <input name="max_marks" type="hidden" value="500">
  <label>Principal Name</label>
  <input name="principal_name" type="text" placeholder="Principal full name" value="{{ s.get('principal_name','Anju Bala') }}" required>
  <button class="btn" type="submit">Generate Certificate</button>
</form>
</div>
</body>
</html>
'''
BULK_DOWNLOAD_HTML = '''
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta http-equiv="refresh" content="2;url=/student">
<title>Downloading</title>
</head>
<body>
<p>Downloading certificate… Next student loading.</p>
<iframe src="/download-pdf" style="display:none;"></iframe>
</body>
</html>
'''
HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>Character Certificate Generator</title>
  <style>
    :root{
      --card:#ffffff;
      --muted:#64748b;
      --accent:#8b5cf6;
    }
    *{box-sizing:border-box}
    body{
      margin:0;
      font-family: -apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
      background: #0f172a;
      color:#1e293b;
      -webkit-font-smoothing:antialiased;
      -moz-osx-font-smoothing:grayscale;
      padding:40px 20px;
      display:flex;
      justify-content:center;
      min-height:100vh;
    }
    .card{
      width:100%;
      max-width:900px;
      background:var(--card);
      border-radius:16px;
      box-shadow:0 4px 12px rgba(0,0,0,0.08);
      padding:50px;
    }
    .header{
      display:flex;
      justify-content:center;
      flex-direction:column;
      align-items:center;
      gap:12px;
      margin-bottom:40px;
      padding-bottom:30px;
      border-bottom:3px solid var(--accent);
    }
    h1{font-size:36px; margin:0; font-weight:900; color:var(--accent); letter-spacing:-1px;}
    .subtitle{color:var(--muted); font-size:16px; font-weight:500; text-align:center}
    form{display:grid; gap:20px;}
    .section{
      background:#f8fafc;
      border-radius:20px;
      padding:28px;
      border:2px solid #e2e8f0;
      transition:border-color 0.2s;
      margin-bottom:24px;
    }
    .section:hover{
      border-color:var(--accent);
    }
    .section h3{margin:0 0 20px 0; font-size:18px; font-weight:800; color:#0f172a; text-transform:uppercase;
     letter-spacing:0.5px; display:flex; align-items:center; gap:10px}
    .section h3::before{content:""; width:4px; height:24px; background:var(--accent); border-radius:2px}
    .grid{
      display:grid;
      grid-template-columns: repeat(2, 1fr);
      gap:16px;
    }
    .grid-col-3{
      display:grid;
      grid-template-columns: repeat(3, 1fr);
      gap:16px;
    }
    label{display:block; font-size:14px; color:#334155; margin-bottom:8px; font-weight:700; text-transform:uppercase; letter-spacing:0.3px}
    input[type="text"], input[type="number"], select {
      width:100%;
      padding:14px 16px;
      border:2px solid #e2e8f0;
      border-radius:10px;
      font-size:15px;
      outline:none;
      background:#fff;
      transition:border-color 0.15s, box-shadow 0.15s;
      font-family:inherit;
    }
    input::placeholder, select::placeholder{color:#cbd5e1}
    input:hover, select:hover { border-color:#cbd5e1; }
    input:focus, select:focus{ 
      border-color:var(--accent); 
      box-shadow:0 0 0 3px rgba(139,92,246,0.1);
      background:#f8fafc;
    }
    input[readonly]{
      background:#f1f5f9;
      color:#64748b;
      cursor:not-allowed;
    }
    .preview{ display:none; }
    .actions{ display:flex; gap:12px; margin-top:30px; flex-wrap:wrap; }
    .btn{
      padding:12px 24px;
      border-radius:10px;
      border:0;
      font-weight:800;
      cursor:pointer;
      transition:background 0.2s;
      font-size:13px;
      text-transform:uppercase;
      letter-spacing:0.6px;
      flex:1;
      min-width:120px;
    }
    .btn-primary{ 
      background:var(--accent);
      color:#fff;
    }
    .btn-primary:hover{ 
      opacity:0.9;
    }
    .btn-ghost{ 
      background:#f1f5f9; 
      border:2px solid #e2e8f0; 
      color:#0f172a;
    }
    .btn-ghost:hover{
      background:var(--accent);
      border-color:var(--accent);
      color:#fff;
    }
    .muted{ color:var(--muted); font-size:13px }
    .helper-text{ display:none; }
    #statusArea{ display:none; }
    .credit{
      font-size:13px;
      margin-top:18px;
      padding-top:14px;
      border-top:2px solid #e2e8f0;
      text-align:center;
      font-weight:700;
      letter-spacing:1.2px;
      padding:12px 24px;
      border-radius:10px;
      background:linear-gradient(135deg,#7c3aed 0%,#a855f7 25%,#d946ef 50%,#ec4899 75%,#f43f5e 100%);
      background-size:200% 200%;
      -webkit-background-clip:text;
      -webkit-text-fill-color:transparent;
      background-clip:text;
      position:relative;
      transition:all 0.35s cubic-bezier(0.34,1.56,0.64,1);
      box-shadow:0 2px 8px rgba(139,92,246,0.15),0 0 12px rgba(217,70,239,0.08),inset 0 1px 2px rgba(255,255,255,0.15);
      display:inline-block;
    }
    .credit::before{
      content:'⭐';
      position:absolute;
      left:2px;
      opacity:0;
      transition:all 0.35s cubic-bezier(0.34,1.56,0.64,1);
      transform:translateX(-12px) scale(0.8);
      font-size:14px;
    }
    .credit::after{
      content:'⭐';
      position:absolute;
      right:2px;
      opacity:0;
      transition:all 0.35s cubic-bezier(0.34,1.56,0.64,1);
      transform:translateX(12px) scale(0.8);
      font-size:14px;
    }
    .credit:hover{
      transform:translateY(-4px) scale(1.04) rotateX(2deg);
      background-position:100% 100%;
      box-shadow:0 8px 20px rgba(139,92,246,0.35),0 0 30px rgba(217,70,239,0.25),0 0 40px 
      rgba(244,63,94,0.1),inset 0 1px 2px rgba(255,255,255,0.25);
    }
    .credit:hover::before{
      opacity:1;
      transform:translateX(-18px) scale(1);
      animation:starFloat 0.6s ease-out;
    }
    .credit:hover::after{
      opacity:1;
      transform:translateX(18px) scale(1);
      animation:starFloat 0.6s ease-out;
    }
    @keyframes starFloat{
      0%{transform:translateX(0) scale(0.8) translateY(0)}
      50%{transform:translateX(0) scale(1.1) translateY(-4px)}
      100%{transform:translateX(0) scale(1) translateY(0)}
    }
    .credit a{ color:var(--accent); text-decoration:none; }
    .credit a:hover{ text-decoration:underline; }
    @media (max-width:1024px){
      body{ padding:24px 16px; }
      .card{ padding:40px; }
      h1{ font-size:32px; }
    }
    @media (max-width:768px){
      body{ padding:20px 12px; }
      .card{ padding:32px 20px; }
      h1{ font-size:28px; }
      .section{ padding:20px; margin-bottom:16px; }
      .grid, .grid-col-3{ gap:12px; }
      label{ font-size:13px; margin-bottom:6px; }
      input[type="text"], input[type="number"], select{ padding:12px 14px; font-size:14px; }
      .btn{ padding:12px 16px; font-size:13px; min-width:100px; }
      .actions{ gap:8px; }
    }
    @media (max-width:640px){
      body{ padding:16px 12px; }
      .card{ padding:20px 16px; border-radius:12px; }
      h1{ font-size:24px; }
      .header{ margin-bottom:24px; padding-bottom:20px; border-bottom-width:2px; }
      .subtitle{ font-size:14px; }
      .section{ padding:16px 14px; margin-bottom:12px; border-radius:12px; }
      .section h3{ font-size:15px; margin-bottom:14px; }
      .grid, .grid-col-3{ grid-template-columns: 1fr; gap:10px; }
      label{ font-size:12px; margin-bottom:6px; margin-top:10px; }
      input[type="text"], input[type="number"], select{ 
        width:100%;
        padding:12px;
        border-radius:8px;
        font-size:16px;
        -webkit-appearance: none;
        appearance: none;
      }
      input:focus, select:focus{ 
        border-color:var(--accent);
        box-shadow: 0 0 0 3px rgba(139,92,246,0.1);
      }
      .btn{ 
        padding:12px 14px;
        font-size:13px;
        min-width:90px;
        flex:1 1 auto;
        touch-action: manipulation;
      }
      .btn:active{ transform: scale(0.96); }
      .actions{ gap:8px; flex-wrap:wrap; }
      .muted{ font-size:12px; }
    }
    @media (max-width:480px){
      body{ padding:12px 8px; -webkit-user-select: none; }
      .card{ padding:16px 12px; border-radius:10px; }
      h1{ font-size:22px; font-weight:900; margin:0; }
      .header{ margin-bottom:16px; padding-bottom:16px; gap:6px; }
      .subtitle{ font-size:12px; }
      .section{ padding:14px 12px; margin-bottom:10px; border-radius:10px; }
      .section h3{ font-size:13px; margin-bottom:12px; gap:6px; }
      .section h3::before{ width:3px; height:18px; }
      .grid, .grid-col-3{ grid-template-columns: 1fr; gap:8px; }
      label{ font-size:11px; margin-bottom:4px; margin-top:8px; text-transform:capitalize; }
      input[type="text"], input[type="number"], select{ 
        width:100%;
        padding:11px;
        border-radius:6px;
        font-size:16px;
        margin-top:4px;
        -webkit-appearance: none;
        appearance: none;
        background-image: url("data:image/svg+xml;charset=UTF-8,%3csvg xmlns='http://www.w3.org/2000/svg' 
        viewBox='0 0 24 24' fill='none' stroke='currentColor' 
        stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3e%3cpolyline points='6 9 12 15 18 9'%3e%3c/polyline%3e%3c/svg%3e");
        background-repeat: no-repeat;
        background-position: right 10px center;
        background-size: 18px;
        padding-right: 32px;
      }
      select{ background-color: #fff; }
      input:focus, select:focus{ border-color:var(--accent); outline:none; box-shadow:0 0 0 2px rgba(139,92,246,0.1); }
      .btn{ 
        padding:8px 10px;
        font-size:11px;
        min-width:65px;
        flex:1 1 auto;
        border-radius:6px;
        font-weight:700;
        touch-action: manipulation;
        -webkit-user-select: none;
      }
      .btn:active{ transform: scale(0.94) !important; }
      .actions{ gap:4px; margin-top:4px; }
      .helper-text{ font-size:11px; margin-top:14px; }
      .muted{ font-size:11px; }
    }
    @media (max-width:360px){
      .card{ padding:12px 10px; }
      h1{ font-size:20px; }
      input[type="text"], input[type="number"], select{ padding:10px; font-size:15px; }
      .btn{ padding:7px 9px; font-size:10px; min-width:60px; }
      label{ font-size:10px; }
    }
  </style>
</head>
<body>
  <div class="card" role="main">
    <div class="header">
      <div>
        <h1>Character Certificate</h1>
        <div class="subtitle">Professional certificate generation tool with modern UI</div>
      </div>
    </div>
    <div>
      <form id="certForm" method="post" action="/generate" novalidate>
        <div class="section" aria-labelledby="student">
          <h3 id="student">Student Details</h3>
          <div class="grid">
            <div>
              <label for="student_name">Student Name</label>
              <input id="student_name" name="student_name" type="text" placeholder="Full name" required>
            </div>
            <div>
              <label for="roll_number">Roll Number</label>
              <input id="roll_number" name="roll_number" type="text" placeholder="Roll number" required>
            </div>
            <div>
              <label for="father_name">Father's Name</label>
              <input id="father_name" name="father_name" type="text" placeholder="Father's name" required>
            </div>
            <div>
              <label for="mother_name">Mother's Name</label>
              <input id="mother_name" name="mother_name" type="text" placeholder="Mother's name" required>
            </div>
            <div>
              <label>Class Last Studied</label>
              <select id="class_studied" name="class_studied" required>
                <option value="" disabled selected>Select class</option>
                <option>X</option><option>XII</option>
              </select>
            </div>
            <div style="grid-column:1 / -1;">
              <label>Subjects</label>
              <div class="grid-col-3">
                <input type="text" name="subject1" placeholder="Subject 1">
                <input type="text" name="subject2" placeholder="Subject 2">
                <input type="text" name="subject3" placeholder="Subject 3">
                <input type="text" name="subject4" placeholder="Subject 4">
                <input type="text" name="subject5" placeholder="Subject 5">
                <input type="text" name="subject6" placeholder="Subject 6">
              </div>
            </div>
            <div class="grid-col-3">
              <div>
                <label for="dob_day">DOB (DD)</label>
                <input id="dob_day" name="dob_day" type="number" min="1" max="31" placeholder="DD" required>
              </div>
              <div>
                <label for="dob_month">DOB (MM)</label>
                <input id="dob_month" name="dob_month" type="number" min="1" max="12" placeholder="MM" required>
              </div>
              <div>
                <label for="dob_year">DOB (YYYY)</label>
                <input id="dob_year" name="dob_year" type="number" min="1900" max="2100" placeholder="YYYY" required>
              </div>
            </div>
            <div class="grid-col-3">
              <div>
                <label for="doa_day">Admission DD</label>
                <input id="doa_day" name="doa_day" type="number" min="1" max="31" placeholder="DD" required>
              </div>
              <div>
                <label for="doa_month">Admission MM</label>
                <input id="doa_month" name="doa_month" type="number" min="1" max="12" placeholder="MM" required>
              </div>
              <div>
                <label for="doa_year">Admission YYYY</label>
                <input id="doa_year" name="doa_year" type="number" min="1900" max="2100" placeholder="YYYY" required>
              </div>
            </div>
          </div>
        </div>
        <div class="section" aria-labelledby="cert">
          <h3 id="cert">Certificate Details</h3>
          <div class="grid-col-3" style="grid-column:1 / -1;">
              <div>
                <label for="issue_day">Issue DD</label>
                <input id="issue_day" name="issue_day" type="number" min="1" max="31" placeholder="DD" required>
              </div>
              <div>
                <label for="issue_month">Issue MM</label>
                <input id="issue_month" name="issue_month" type="number" min="1" max="12" placeholder="MM" required>
              </div>
              <div>
                <label for="issue_year">Issue YYYY</label>
                <input id="issue_year" name="issue_year" type="number" min="1900" max="2100" placeholder="YYYY" required>
              </div>
            </div>
            <div style="grid-column:1 / -1;">
              <label for="principal_name">Principal Name</label>
              <input id="principal_name" name="principal_name" type="text" placeholder="Principal full name" value="Anju Bala" required>
            </div>
          </div>
        </div>
        <div class="section" aria-labelledby="additional">
          <h3 id="additional">Additional Details</h3>
          <div class="grid">
            <div>
              <label for="exam_type">Exam Type (Auto-filled)</label>
              <input id="exam_type" name="exam_type" type="text" placeholder="Auto-filled based on class" readonly required>
            </div>
            <div>
              <label for="exam_year">Exam Year</label>
              <input id="exam_year" name="exam_year" type="number" min="1900" max="2100" placeholder="Year" required>
            </div>
          </div>
          <div class="grid">
            <div>
              <label for="marks_obtained">Marks Obtained</label>
              <input id="marks_obtained" name="marks_obtained" type="number" placeholder="e.g. 85" required>
            </div>
            <div>
              <label>Maximum Marks</label>
              <input type="text" value="500" readonly style="background:#f1f5f9; cursor:not-allowed;">
            </div>
          </div>
          <div style="grid-column:1 / -1;">
            <label for="class_incharge_name">Class Incharge Name</label>
            <input id="class_incharge_name" name="class_incharge_name" type="text" placeholder="Class teacher/incharge name">
          </div>
        </div>
        <input name="max_marks" type="hidden" value="500">
        <input name="dol_day" type="hidden" value="">
        <input name="dol_month" type="hidden" value="">
        <input name="dol_year" type="hidden" value="">
        <div style="display:flex; gap:12px; margin-top:8px;">
          <button class="btn btn-primary" id="generateBtn" type="submit">Generate PDF</button>
          <button class="btn btn-ghost" id="downloadExample" type="button">Example</button>
          <button class="btn btn-ghost" id="bulkBtn" type="button" onclick="window.location='/bulk'">Bulk</button>
          <button class="btn btn-ghost" id="clearBtn" type="button">Clear</button>
        </div>
      </form>
      <div class="helper-text" style="margin-top:24px; text-align:center">
      Fill all required fields and click Generate PDF to download your professional certificate</div>
    </div>
  </div>
  <script>
    (function(){
      const t = new Date();
      const dd = String(t.getDate()).padStart(2,'0');
      const mm = String(t.getMonth()+1).padStart(2,'0');
      const yyyy = t.getFullYear();
      document.getElementById('issue_day').value = dd;
      document.getElementById('issue_month').value = mm;
      document.getElementById('issue_year').value = yyyy;
      document.getElementById('exam_year').value = yyyy;
      const form = document.getElementById('certForm');
      const clearBtn = document.getElementById('clearBtn');
      const classSelect = document.getElementById('class_studied');
      const examTypeInput = document.getElementById('exam_type');
      classSelect.addEventListener('change', function(){
        const selectedClass = this.value;
        if(selectedClass === 'X' || selectedClass === 'IX') {
          examTypeInput.value = 'AISSE';
        } else if(selectedClass === 'XII' || selectedClass === 'XI') {
          examTypeInput.value = 'AISSCE';
        } else {
          examTypeInput.value = 'AISSE';
        }
      });
      clearBtn.addEventListener('click', function(){
        form.reset();
        document.getElementById('issue_day').value = dd;
        document.getElementById('issue_month').value = mm;
        document.getElementById('issue_year').value = yyyy;
        document.getElementById('exam_year').value = yyyy;
        document.getElementById('principal_name').value = 'Anju Bala';
        examTypeInput.value = '';
      });
      document.getElementById('downloadExample').addEventListener('click', function(){
        const example = {
          student_name: "Rahul Kumar",
          father_name: "Ramesh Kumar",
          mother_name: "Sita Devi",
          dob_day: "15", dob_month: "06", dob_year: "2010",
          roll_number: "12345678",
          class_studied: "X",
          doa_day: "01", doa_month: "06", doa_year: "2016",
          subject1: "English", subject2: "Hindi", subject3: "Mathematics",
          subject4: "Science", subject5: "Social Science", subject6: "Physical Education",
          exam_type: "AISSE",
          exam_year: "2024",
          marks_obtained: "425",
          class_incharge_name: "Ms. Priya Sharma",
          issue_day: dd, issue_month: mm, issue_year: yyyy,
          principal_name: "Anju Bala"
        };
        for(const k in example){
          const el = document.getElementsByName(k)[0];
          if(el) el.value = example[k];
        }
        form.submit();
      });
      form.addEventListener('submit', function(e){
        const required = ['student_name','father_name','mother_name',
                          'dob_day','dob_month','dob_year','roll_number','class_studied',
                          'doa_day','doa_month','doa_year',
                          'issue_day','issue_month','issue_year','principal_name',
                          'exam_type','exam_year','marks_obtained'];
        let missing = [];
        required.forEach(function(name){
          const el = document.getElementsByName(name)[0];
          if(!el) return;
          if(String(el.value).trim() === '') missing.push(name);
        });
        if(missing.length){
          e.preventDefault();
          alert('⚠️ Please fill all required fields!');
          return false;
        }
      });
    })();
  </script>
</body>
</html>
"""
def build_pdf(data):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(A4))
    width, height = landscape(A4)
    def get_val(key, default=''):
        val = data.get(key, default)
        if val is None or (isinstance(val, str) and not val.strip()):
            return default
        return str(val).strip()
    LMARGIN = 50
    RMARGIN = 50
    TOP_MARGIN = 40
    BOTTOM_MARGIN = 40
    c.setLineWidth(2.5)
    c.rect(LMARGIN - 15, BOTTOM_MARGIN, width - (LMARGIN * 2) + 30, height - TOP_MARGIN - BOTTOM_MARGIN)
    c.setLineWidth(0.5)
    c.rect(LMARGIN - 12, BOTTOM_MARGIN + 3, width - (LMARGIN * 2) + 24, height - TOP_MARGIN - BOTTOM_MARGIN - 6)
    y = height - TOP_MARGIN - 20
    c.setFont("Times-Bold", 15)
    school_name = "Govt. Model Sanskriti Senior Secondary School, ADAMPUR MANDI"
    c.drawCentredString(width / 2, y, school_name)
    y -= 20
    c.setFont("Times-Bold", 12)
    affiliation = "(Affiliated to CBSE, New Delhi)-44090"
    c.drawCentredString(width / 2, y, affiliation)
    y -= 16
    c.setFont("Times-Roman", 11)
    school_addr = "Opposite Canara Bank, Adampur Mandi (Hisar)"
    c.drawCentredString(width / 2, y, school_addr)
    y -= 25
    c.setLineWidth(1)
    c.line(LMARGIN + 20, y, width - LMARGIN - 20, y)
    y -= 20
    c.setFont("Times-Bold", 18)
    c.drawCentredString(width / 2, y, "CHARACTER CERTIFICATE")
    y -= 30
    c.setFont("Times-Roman", 11)
    c.drawString(LMARGIN + 20, y, "Date: [")
    issue_date = f"{get_val('issue_day','__')}/{get_val('issue_month','__')}/{get_val('issue_year','____')}"
    c.setLineWidth(1)
    c.line(LMARGIN + 60, y - 2, LMARGIN + 140, y - 2)
    c.drawString(LMARGIN + 145, y, "]")
    c.setFont("Times-Roman", 11)
    c.drawCentredString(LMARGIN + 100, y - 1, issue_date)
    y -= 20
    c.setFont("Times-Roman", 11)
    student_name = get_val("student_name", "").upper()
    father_name = get_val("father_name", "")
    mother_name = get_val("mother_name", "")
    para = "This is to certify that "
    c.drawString(LMARGIN + 20, y, para)
    x_pos = LMARGIN + 20 + c.stringWidth(para)
    c.setLineWidth(1)
    c.line(x_pos, y - 2, x_pos + 260, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos + 130, y - 1, student_name[:32] if student_name else "")
    c.setFont("Times-Roman", 11)
    para2 = " S/D/O Shri "
    c.drawString(x_pos + 270, y, para2)
    x_pos2 = x_pos + 270 + c.stringWidth(para2)
    c.setLineWidth(1)
    c.line(x_pos2, y - 2, x_pos2 + 220, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos2 + 110, y - 1, father_name[:25] if father_name else "")
    c.setFont("Times-Roman", 11)
    c.drawString(x_pos2 + 230, y, " and")
    y -= 18
    para3 = "Shrimati "
    c.drawString(LMARGIN + 20, y, para3)
    x_pos = LMARGIN + 20 + c.stringWidth(para3)
    c.setLineWidth(1)
    c.line(x_pos, y - 2, x_pos + 220, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos + 110, y - 1, mother_name[:25] if mother_name else "")
    c.setFont("Times-Roman", 11)
    c.drawString(x_pos + 230, y, " was a bonafide student of this institution from ")
    c.setFont("Times-Roman", 11)
    doa = f"{get_val('doa_day','__')}/{get_val('doa_month','__')}/{get_val('doa_year','____')}"
    x_pos2 = x_pos + 230 + c.stringWidth(" was a bonafide student of this institution from ")
    c.setLineWidth(1)
    c.line(x_pos2, y - 2, x_pos2 + 95, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos2 + 47.5, y - 1, doa)
    c.setFont("Times-Roman", 11)
    c.drawString(x_pos2 + 105, y, " to ")
    dol = f"{get_val('dol_day','__')}/{get_val('dol_month','__')}/{get_val('dol_year','____')}"
    x_pos3 = x_pos2 + 105 + c.stringWidth(" to ")
    c.setLineWidth(1)
    c.line(x_pos3, y - 2, x_pos3 + 95, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos3 + 47.5, y - 1, dol)
    c.setFont("Times-Roman", 11)
    c.drawString(x_pos3 + 105, y, ".")
    y -= 22
    class_studied = get_val("class_studied", "X/XII")
    exam_type = get_val("exam_type", "")
    exam_year = get_val("exam_year", "")
    exam_text = f"He/She appeared in the Class {class_studied} (AISSE/AISSCE) "
    c.drawString(LMARGIN + 20, y, exam_text)
    x_pos = LMARGIN + 20 + c.stringWidth(exam_text)
    c.setLineWidth(1)
    c.line(x_pos, y - 2, x_pos + 95, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos + 47.5, y - 1, exam_type[:15] if exam_type else "")
    c.setFont("Times-Roman", 11)
    c.drawString(x_pos + 105, y, " CBSE Board Examination held in ")
    x_pos2 = x_pos + 105 + c.stringWidth(" CBSE Board Examination held in ")
    c.setLineWidth(1)
    c.line(x_pos2, y - 2, x_pos2 + 95, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos2 + 47.5, y - 1, str(exam_year)[:15] if exam_year else "")
    c.setFont("Times-Roman", 11)
    c.drawString(x_pos2 + 105, y, ",")
    y -= 20
    roll_number = get_val("roll_number", "")
    marks_obtained = get_val("marks_obtained", "")
    max_marks = get_val("max_marks", "")
    roll_text = "bearing Roll Number: ["
    c.drawString(LMARGIN + 20, y, roll_text)
    x_pos = LMARGIN + 20 + c.stringWidth(roll_text)
    c.setLineWidth(1)
    c.line(x_pos, y - 2, x_pos + 130, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos + 65, y - 1, str(roll_number)[:28] if roll_number else "")
    c.setFont("Times-Roman", 11)
    c.drawString(x_pos + 140, y, "] securing (Marks Obtained) ")
    x_pos2 = x_pos + 140 + c.stringWidth("] securing (Marks Obtained) ")
    c.setLineWidth(1)
    c.line(x_pos2, y - 2, x_pos2 + 85, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos2 + 42.5, y - 1, str(marks_obtained) if marks_obtained else "")
    c.setFont("Times-Roman", 11)
    c.drawString(x_pos2 + 95, y, " Out of (Maximum Marks) ")
    x_pos3 = x_pos2 + 95 + c.stringWidth(" Out of (Maximum Marks) ")
    c.setLineWidth(1)
    c.line(x_pos3, y - 2, x_pos3 + 85, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos3 + 42.5, y - 1, str(max_marks) if max_marks else "")
    c.setFont("Times-Roman", 11)
    c.drawString(x_pos3 + 95, y, ".")
    y -= 20
    dob_day = get_val('dob_day', '__')
    dob_month = get_val('dob_month', '__')
    dob_year = get_val('dob_year', '____')
    dob = f"{dob_day}/{dob_month}/{dob_year}"
    dob_text = "The Date of Birth (DOB) as officially registered in school record is "
    c.drawString(LMARGIN + 20, y, dob_text)
    x_pos = LMARGIN + 20 + c.stringWidth(dob_text)
    c.setLineWidth(1)
    c.line(x_pos, y - 2, x_pos + 110, y - 2)
    c.setFont("Times-Roman", 11)
    c.drawCentredString(x_pos + 55, y - 1, dob)
    c.setFont("Times-Roman", 11)
    c.drawString(x_pos + 120, y, ".")
    y -= 22
    subjects_text = "During his/her tenure at this institution, he/she studied the following subjects:"
    c.drawString(LMARGIN + 20, y, subjects_text)
    y -= 18
    c.setFont("Times-Roman", 10)
    box_width = 145
    box_height = 22
    spacing_x = 165
    spacing_y = 28
    for i in range(1, 7):
        subj = get_val(f"subject{i}", "").strip()
        col = (i - 1) % 3
        row = (i - 1) // 3
        x_box = LMARGIN + 20 + col * spacing_x
        y_box = y - row * spacing_y
        c.setLineWidth(1)
        c.rect(x_box, y_box - box_height, box_width, box_height)
        if subj:
            c.setFont("Times-Roman", 11)
            c.drawCentredString(x_box + box_width/2, y_box - 16, subj[:18])
    y -= 60
    c.setLineWidth(1)
    c.line(LMARGIN + 20, y, width - LMARGIN - 20, y)
    y -= 18
    c.setFont("Times-Roman", 11)
    char_line1 = "He/She character and conduct have been good throughout the period of study. There is no"
    c.drawString(LMARGIN + 20, y, char_line1)
    y -= 16
    char_line2 = "disciplinary action or adverse remark on record."
    c.drawString(LMARGIN + 20, y, char_line2)
    y -= 24
    c.setLineWidth(1)
    c.line(LMARGIN + 20, y, width - LMARGIN - 20, y)
    y -= 20
    c.setFont("Times-Roman", 11)
    class_incharge = get_val("class_incharge_name", "")
    principal = get_val("principal_name", "")
    sig_x_left = LMARGIN + 40
    sig_x_right = width - LMARGIN - 200;
    c.setLineWidth(1)
    c.line(sig_x_left, y - 3, sig_x_left + 180, y - 3)
    c.setFont("Times-Roman", 11)
    if class_incharge:
        c.drawCentredString(sig_x_left + 90, y - 1, class_incharge[:25])
    c.setFont("Times-Roman", 10)
    c.drawString(sig_x_left, y - 18, "Class Incharge")
    c.setLineWidth(1)
    c.line(sig_x_right, y - 3, sig_x_right + 180, y - 3)
    c.setFont("Times-Roman", 11)
    if principal:
        c.drawCentredString(sig_x_right + 90, y - 1, principal[:25])
    c.setFont("Times-Roman", 10)
    c.drawString(sig_x_right + 50, y - 18, "Principal")
    c.save()
    buf.seek(0)
    return buf
@app.route("/")
def index():
    t = datetime.now()
    return render_template_string(HTML, d=f"{t.day:02d}", m=f"{t.month:02d}", y=t.year)
@app.route('/bulk', methods=['GET', 'POST'])
def bulk_upload():
  if request.method == 'POST':
    try:
      wb = load_workbook(request.files['excel'], data_only=True)
      ws = wb[wb.sheetnames[0]]
      headers = [clean(c.value) for c in ws[1]]
      students = []
      for row in ws.iter_rows(min_row=2, values_only=True):
        raw = {headers[i]: clean(row[i]) if i < len(row) else '' for i in range(len(headers))}
        student_name = ''
        father_name = ''
        mother_name = ''
        dob = ''
        cls = ''
        admission_no = ''
        admission_date = ''
        subjects = []
        exam_type = ''
        exam_year = ''
        marks_obtained = ''
        principal_name = 'Anju Bala'
        def _clean_person(val):
          if not val:
            return ''
          v = str(val).strip()
          if v.lower() in ('yes','no','y','n','true','false'):
            return ''
          return v
        father_candidates = []
        mother_candidates = []
        norm_map = {}
        for h in headers:
          if not h:
            continue
          lh = h.lower()
          lh_n = re.sub(r"[^a-z0-9 ]", "", lh).strip()
          norm_map[lh_n] = h
        for nh, orig in norm_map.items():
          if (not father_name) and (('father' in nh and 'full' in nh and 'name' in nh) or ('father' in nh and 'aadhar' in nh)):
            father_name = _clean_person(raw.get(orig, ''))
          if (not mother_name) and (('mother' in nh and 'full' in nh and 'name' in nh) or ('mother' in nh and 'aadhar' in nh)):
            mother_name = _clean_person(raw.get(orig, ''))
        for h in headers:
          if not h:
            continue
          lh = h.lower()
          if (not father_name) and ('father' in lh and 'full' in lh and 'name' in lh):
            father_name = _clean_person(raw.get(h, ''))
          if (not mother_name) and ('mother' in lh and 'full' in lh and 'name' in lh):
            mother_name = _clean_person(raw.get(h, ''))
        exact_father_headers = ["Father's Full Name aso on Aadhar Card","Father's Full Name as on Aadhar Card",
                                "Father's Full Name as on Aadhaar","Father's Full Name as on Aadhar","Father's Full Name as on Aadhar Card"]
        exact_mother_headers = ["Mother's Full Name as on Aadhaar","Mother's Full Name as on Aadhar Card","Mother's Full Name as on Aadhar"]
        for eh in exact_father_headers:
          if not father_name and eh in raw and raw.get(eh):
            father_name = _clean_person(raw.get(eh))
        for eh in exact_mother_headers:
          if not mother_name and eh in raw and raw.get(eh):
            mother_name = _clean_person(raw.get(eh))
        OCCUPATION_KEYS = ('occupation','job','profession','designation','post','work')
        EDUCATION_KEYS = ('education','qualification','degree','edu','qualification','qualification details','qualification_level',
                          'qualificaton','highest qual')
        for h in headers:
          if not h:
            continue
          lh = h.lower()
          lh_n = re.sub(r"[^a-z0-9 ]", "", lh).strip()
          val = raw.get(h, '')
          if lh_n == "mothers full name as on aadhaar":
            mother_candidates.append((lh_n, val))
          if lh_n == "fathers full name as on aadhaar" or lh_n == "fathers full name aso on aadhaar":
            father_candidates.append((lh_n, val))
          if lh_n == "admission number" or lh_n == "admission_no":
            admission_no = val
          if lh_n == "admission date":
            admission_date = val
          if not student_name and ('full' in lh and 'name' in lh):
            student_name = val
          if 'father' in lh:
            father_candidates.append((lh_n, val))
          if 'mother' in lh:
            mother_candidates.append((lh_n, val))
          if not dob and 'date' in lh and 'birth' in lh:
            if val:
              dob = val
          if not admission_no and 'admission' in lh and ('no' in lh or 'number' in lh or 'admn' in lh):
            admission_no = val
          if not admission_date and 'admission' in lh and 'date' in lh:
            admission_date = val
          if not cls and 'class' in lh:
            cls = val
          if 'subject' in lh and val:
            subjects.append(val)
          if not exam_type and 'exam' in lh and ('type' in lh or 'code' in lh):
            exam_type = val
          if not exam_year and 'exam' in lh and 'year' in lh:
            exam_year = val
          if (not marks_obtained) and 'marks' in lh and ('obtained' in lh or 'scored' in lh or 'marks_obtained' in lh):
            marks_obtained = val
          if not principal_name and 'principal' in lh:
            principal_name = val or 'Anju Bala'
        def choose_person(cands):
          for hdr, c in cands:
            pc = _clean_person(c)
            if not pc:
              continue
            if ('full' in hdr and 'name' in hdr) or ('aadhar' in hdr):
              return pc
          best = ''
          best_score = 0
          for (hdr, c) in cands:
            pc = _clean_person(c)
            if not pc:
              continue
            if any(k in hdr for k in OCCUPATION_KEYS) or any(k in hdr for k in EDUCATION_KEYS):
              continue
            low = pc.lower()
            if any(k in low for k in OCCUPATION_KEYS) or any(k in low for k in EDUCATION_KEYS):
              continue
            if any(ch.isdigit() for ch in pc) or len(pc) < 3:
              continue
            score = len(pc) + (10 if ' ' in pc else 0) + (20 if 'name' in hdr else 0)
            if score > best_score:
              best = pc
              best_score = score
          return best
        if not father_name:
          father_name = choose_person(father_candidates)
        if not mother_name:
          mother_name = choose_person(mother_candidates)
        today = datetime.now()
        def split_date(s):
          if not s:
            return ('','','')
          s_str = str(s).strip()
          if not s_str:
            return ('','','')
          if '/' in s_str:
            parts = [p.strip() for p in s_str.split('/') if p.strip()]
            if len(parts) >= 3:
              return (parts[0], parts[1], parts[2])
            if len(parts) == 2:
              return (parts[0], parts[1], '')
            if len(parts) == 1:
              return (parts[0], '', '')
          if '-' in s_str and len(s_str.split('-')) == 3:
            parts = [p.strip() for p in s_str.split('-') if p.strip()]
            if len(parts) == 3:
              if len(parts[0]) == 4:
                return (parts[2], parts[1], parts[0])
              else:
                return (parts[0], parts[1], parts[2])
          try:
            from datetime import datetime as dt
            for fmt in ['%b %d, %Y', '%B %d, %Y', '%d %b %Y', '%d %B %Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
              try:
                parsed = dt.strptime(s_str, fmt)
                return (f"{parsed.day:02d}", f"{parsed.month:02d}", str(parsed.year))
              except:
                continue
          except:
            pass
          return ('','','')
        dob_d, dob_m, dob_y = split_date(dob)
        doa_d, doa_m, doa_y = split_date(admission_date)
        if student_name:
          stud = {
            'student_name': student_name,
            'father_name': father_name or '',
            'mother_name': mother_name or '',
            'admission_no': admission_no or '',
            'dob': dob or '',
            'dob_day': dob_d or '', 
            'dob_month': dob_m or '', 
            'dob_year': dob_y or '',
            'class_studied': cls or '',
            'doa_day': doa_d or '', 
            'doa_month': doa_m or '', 
            'doa_year': doa_y or '',
            'dol_day': '', 
            'dol_month': '', 
            'dol_year': '',
            'subject1': subjects[0] if len(subjects) > 0 else '',
            'subject2': subjects[1] if len(subjects) > 1 else '',
            'subject3': subjects[2] if len(subjects) > 2 else '',
            'subject4': subjects[3] if len(subjects) > 3 else '',
            'subject5': subjects[4] if len(subjects) > 4 else '',
            'subject6': subjects[5] if len(subjects) > 5 else '',
            'exam_type': exam_type or '',
            'exam_year': exam_year or '',
            'marks_obtained': marks_obtained or '',
            'max_marks': '500',
            'principal_name': principal_name or 'Anju Bala'
          }
          students.append(stud)
      session.clear()
      session['students'] = students
      session['index'] = 0
      session.modified = True
      return redirect('/set-incharge')
    except Exception as _e:
      import traceback
      with open('flask_err.txt','a',encoding='utf-8') as f:
        traceback.print_exc(file=f)
        f.write('\n---\n')
      return abort(500, 'Internal server error (details written to flask_err.txt)')
  return BULK_UPLOAD_HTML
@app.route('/set-incharge', methods=['GET', 'POST'])
def set_incharge():
  if 'students' not in session:
    return redirect('/bulk')
  if request.method == 'POST':
    incharge_name = request.form.get('class_incharge_name', '').strip()
    session['global_class_incharge'] = incharge_name
    session.modified = True
    return redirect('/student')
  count = len(session.get('students', []))
  return render_template_string(CLASS_INCHARGE_FORM, count=count)
@app.route('/student')
def student():
  if 'students' not in session:
    return redirect('/bulk')
  i = session.get('index', 0)
  students = session['students']
  if i >= len(students):
    return '<h2>All certificates generated.</h2><p><a href="/bulk">Upload another Excel file</a></p>'
  raw = students[i].copy()
  global_incharge = session.get('global_class_incharge', '')
  raw['class_incharge_name'] = global_incharge
  keys = ['student_name','father_name','mother_name','admission_no','class_studied',
          'doa_day','doa_month','doa_year',
          'subject1','subject2','subject3','subject4','subject5','subject6',
          'exam_type','exam_year','class_incharge_name','principal_name','marks_obtained']
  s = {k: raw.get(k, '') for k in keys}
  s['dob_day'] = raw.get('dob_day','') or ''
  s['dob_month'] = raw.get('dob_month','') or ''
  s['dob_year'] = raw.get('dob_year','') or ''
  if s['dob_day'] and s['dob_month'] and s['dob_year']:
    s['dob_display'] = f"{s['dob_day']}/{s['dob_month']}/{s['dob_year']}"
  else:
    s['dob_display'] = 'Not available'
  dob = raw.get('dob','') or ''
  s['dob'] = dob
  cls = (s.get('class_studied') or '').upper()
  if not s.get('exam_type'):
    s['exam_type'] = 'AISSE' if cls in ['X','10'] else 'AISSCE'
  if not s.get('exam_year'):
    s['exam_year'] = datetime.now().year
  if not s.get('principal_name'):
    s['principal_name'] = 'Anju Bala'
  return render_template_string(BULK_FORM_HTML, s=s)
@app.route('/generate-one', methods=['POST'])
def generate_one():
  try:
    if 'students' not in session:
      return redirect('/bulk')
    i = session['index']
    students = session['students']
    raw = students[i]
    def val(name, default=''):
      form_val = request.form.get(name)
      if form_val and form_val.strip():
        return form_val.strip()
      raw_val = raw.get(name)
      if raw_val and str(raw_val).strip():
        return str(raw_val).strip()
      return default
    cls = (val('class_studied') or '').upper()
    today = datetime.now()
    dob_day = val('dob_day', '')
    dob_month = val('dob_month', '')
    dob_year = val('dob_year', '')
    doa_day = val('doa_day', '')
    doa_month = val('doa_month', '')
    doa_year = val('doa_year', '')
    exam_type = val('exam_type', '')
    if not exam_type:
      exam_type = 'AISSE' if cls in ['X','10'] else 'AISSCE'
    exam_year = val('exam_year', '')
    if not exam_year:
      exam_year = str(today.year)
    principal_name = val('principal_name', '')
    if not principal_name:
      principal_name = 'Anju Bala'
    class_incharge_name = session.get('global_class_incharge', '')
    roll_number = request.form.get('roll_number', '').strip()
    data = {
      'student_name': val('student_name', ''),
      'father_name': raw.get('father_name', ''),
      'mother_name': raw.get('mother_name', ''),
      'admission_no': raw.get('admission_no', ''),
      'roll_number': roll_number,
      'dob_day': dob_day,
      'dob_month': dob_month,
      'dob_year': dob_year,
      'class_studied': val('class_studied', ''),
      'doa_day': doa_day,
      'doa_month': doa_month,
      'doa_year': doa_year,
      'exam_type': exam_type,
      'exam_year': exam_year,
      'marks_obtained': val('marks_obtained', ''),
      'max_marks': '500',
      'dol_day': f"{today.day:02d}",
      'dol_month': f"{today.month:02d}",
      'dol_year': f"{today.year}",
      'subject1': request.form.get('subject1','').strip(),
      'subject2': request.form.get('subject2','').strip(),
      'subject3': request.form.get('subject3','').strip(),
      'subject4': request.form.get('subject4','').strip(),
      'subject5': request.form.get('subject5','').strip(),
      'subject6': request.form.get('subject6','').strip(),
      'class_incharge_name': class_incharge_name,
      'issue_day': f"{today.day:02d}",
      'issue_month': f"{today.month:02d}",
      'issue_year': f"{today.year}",
      'principal_name': principal_name
    }
    pdf_buf = build_pdf(data)
    pdf_bytes = pdf_buf.getvalue()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    try:
      tmp.write(pdf_bytes)
      tmp.flush()
      tmp_path = tmp.name
    finally:
      tmp.close()
    key = str(uuid.uuid4())
    DATA_STORE[key] = tmp_path
    session['last_pdf_key'] = key
    session['last_name'] = data.get('student_name','certificate')
    session['index'] += 1
    session.modified = True
    return BULK_DOWNLOAD_HTML
  except Exception as exc:
    import traceback
    traceback.print_exc()
    raise
@app.route('/download-pdf')
def download_pdf():
  key = session.get('last_pdf_key')
  if not key:
    return redirect('/bulk')
  path = DATA_STORE.pop(key, None)
  if not path or not os.path.exists(path):
    return redirect('/bulk')
  with open(path, 'rb') as f:
    data = f.read()
  try:
    os.unlink(path)
  except Exception:
    pass
  buf = io.BytesIO(data)
  name = session.get('last_name', 'certificate')
  safe = re.sub(r"[^\w]", "_", name)
  try:
    return send_file(buf, as_attachment=True, download_name=f"Character_Certificate_{safe}.pdf", mimetype='application/pdf')
  except TypeError:
    return send_file(buf, as_attachment=True, attachment_filename=f"Character_Certificate_{safe}.pdf", mimetype='application/pdf')
@app.route('/debug-students')
def debug_students():
  import json
  if 'students' not in session:
    return 'NO_STUDENTS'
  i = session.get('index', 0)
  students = session.get('students', [])
  if i >= len(students):
    return 'INDEX_OUT_OF_RANGE'
  return json.dumps(students[i], ensure_ascii=False)
@app.route("/generate", methods=["POST"])
def generate():
    required = ["student_name","father_name","mother_name",
          "dob_day","dob_month","dob_year","roll_number","class_studied",
          "doa_day","doa_month","doa_year",
          "issue_day","issue_month","issue_year","principal_name",
          "exam_type","exam_year","marks_obtained"]
    for r in required:
        if not request.form.get(r):
            return abort(400, f"Missing required field: {r}")
    data = {k: request.form.get(k, "") for k in request.form.keys() if k not in ["school_name", "school_code", "school_address"]}
    date_fields = {"dob_day": (1, 31), "dob_month": (1, 12), "dob_year": (1900, 2100),
                   "doa_day": (1, 31), "doa_month": (1, 12), "doa_year": (1900, 2100),
                   "issue_day": (1, 31), "issue_month": (1, 12), "issue_year": (1900, 2100)}
    for field, (min_val, max_val) in date_fields.items():
        try:
            val = int(data[field])
            if not (min_val <= val <= max_val):
                return abort(400, f"Invalid {field}: must be between {min_val} and {max_val}")
        except (ValueError, KeyError):
            return abort(400, f"Invalid {field}: must be a valid number")
    now = datetime.now()
    data['principal_name'] = 'Anju Bala'
    data['issue_day'] = f"{now.day:02d}"
    data['issue_month'] = f"{now.month:02d}"
    data['issue_year'] = f"{now.year}"
    data['max_marks'] = '500'
    data['dol_day'] = f"{now.day:02d}"
    data['dol_month'] = f"{now.month:02d}"
    data['dol_year'] = f"{now.year}"
    pdf_buf = build_pdf(data)
    raw_name = (data.get("student_name") or "certificate").strip()
    safe = re.sub(r"[^\w]", "_", raw_name)
    try:
        return send_file(pdf_buf, as_attachment=True, download_name=f"Character_Certificate_{safe}.pdf", mimetype='application/pdf')
    except TypeError:
        return send_file(pdf_buf, as_attachment=True, attachment_filename=f"Character_Certificate_{safe}.pdf", mimetype='application/pdf')
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)